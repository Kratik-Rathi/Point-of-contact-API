from flask import Flask, Response, jsonify, request
import pyodbc
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

app = Flask(__name__)

# Database Connection Credentials
DB_SERVER = '10.0.0.180'
DB_NAME = 'LuddyHack'
DB_USERNAME = 'testing'
DB_PASSWORD = 'password'
DB_DRIVER = '{ODBC Driver 17 for SQL Server}'
DB_TRUSTED_CONNECTION = 'no'
DB_TRUST_SERVER_CERTIFICATE = 'yes'

# Database Connection String
conn_str = (
    f"DRIVER={DB_DRIVER};"
    f"SERVER={DB_SERVER};"
    f"DATABASE={DB_NAME};"
    f"UID={DB_USERNAME};"
    f"PWD={DB_PASSWORD};"
    f"Trusted_Connection={DB_TRUSTED_CONNECTION};"
    f"TrustServerCertificate={DB_TRUST_SERVER_CERTIFICATE};"
)

# Email Configuration
EMAIL_SERVER = 'smtp.gmail.com'
EMAIL_PORT = 587
EMAIL_USERNAME = 'mail id'
EMAIL_PASSWORD = 'api password'


def get_db_connection():
    try:
        connection = pyodbc.connect(conn_str)
        return connection
    except pyodbc.Error as e:
        print(f"Database connection failed: {e}")
        raise Exception("Please check database settings.")


def execute_query_and_fetch_data(query, params=None):
    try:
        # The database connection
        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(query, params or ())
        result_data = cursor.fetchall()
        columns = [i[0] for i in cursor.description]
        cursor.close()
        connection.close()
        return result_data, columns
    except Exception as e:
        print(f"Error executing query: {e}")
        raise Exception("Failed to execute database query.")


def send_mail(merged_excel_file, mail_subject, mail_body, receivers, ccs):
    try:
        message = MIMEMultipart()
        message['Subject'] = mail_subject
        message['From'] = EMAIL_USERNAME
        message['To'] = ', '.join(receivers)
        message['CC'] = ', '.join(ccs)
        message.attach(MIMEText(mail_body, 'plain'))
        with open(merged_excel_file, 'rb') as attachment:
            part = MIMEApplication(attachment.read(), Name=f'{merged_excel_file}_report.xlsx')
            part['Content-Disposition'] = f'attachment; filename={merged_excel_file}_report.xlsx'
            message.attach(part)

        with smtplib.SMTP(EMAIL_SERVER, EMAIL_PORT) as smtpserver:
            smtpserver.starttls()  # Secure the connection
            smtpserver.login(EMAIL_USERNAME, EMAIL_PASSWORD)
            all_receiver_emails = receivers + ccs
            smtpserver.sendmail(EMAIL_USERNAME, all_receiver_emails, message.as_string())
        print("Email sent successfully")
    except smtplib.SMTPException as e:
        print(f"Error sending email: {e}")
        raise Exception("Failed to send email. Please check your email settings.")


def convert_to_dataframe(result_data, columns):
    try:
        result_df = pd.DataFrame(columns=columns)
        if result_data:
            result_data = list(map(list, zip(*result_data)))
            result_df = pd.DataFrame(dict(zip(columns, result_data)))
        result_df.fillna(0, inplace=True)
        return result_df
    except Exception as e:
        print(f"Error converting to DataFrame: {e}")
        raise Exception("Failed to convert query result to DataFrame.")


@app.route('/api/poc', methods=['GET'])
def point_of_contact_route():
    try:
        product_name = request.args.get('product_name', default=None, type=str)
        repository_name = request.args.get('repository_name', default=None, type=str)
        query = "EXECUTE FindPOC"
        params = []
        # Checking which parameter is provided (either product_name or repository_name)
        if product_name:
            query += " @ProductName = ?"
            params.append(product_name)
        elif repository_name:
            query += " @RepositoryName = ?"
            params.append(repository_name)
        else:
            raise Exception("Neither product_name nor repository_name was provided.")
        result_data, columns = execute_query_and_fetch_data(query, params)

        poc_df = convert_to_dataframe(result_data, columns)
        poc_df.to_excel('point_of_contact.xlsx', index=False, header=True)
        # Opening the Excel file and apply formatting
        wb = load_workbook('point_of_contact.xlsx')
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
                cell.border = border
        # Adjusting column widths based on content
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    print(f"An exception occurred: {e}")
            adjusted_width = max_length + 2  # Add some padding
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        wb.save('point_of_contact.xlsx')
        # Emailing with the attachment
        mail_subject = f"Point of Contacts {datetime.now().strftime('%Y-%m-%d')}"
        mail_body = "Please find attached the point of contacts for the project."
        receiver_emails = ['receiver mail(s)']
        cc_emails = ['cc mail (s)']
        send_mail("point_of_contact.xlsx", mail_subject, mail_body, receiver_emails, cc_emails,)

        return Response("Contact Generated and Email Sent Successfully!")

    except Exception as e:
        print(f"Error occurred: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
